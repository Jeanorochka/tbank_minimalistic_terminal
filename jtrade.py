# - выбор брокерского счета при запуске;
# - открытие LONG/SHORT лимиткой по текущей лучшей цене стакана;
# - SL% меняется прямо внутри окна;
# - TP% считается автоматически по RR 1:2;
# - по умолчанию SL = 1.3%, TP = 2.6% по RR 1:2;
# - ручное закрытие LONG/SHORT только рыночной заявкой;
# - автожурнал сделок в Excel: Дата / Капитал счёта / PnL / Тикер / Комментарии;
# - портфель внутри интерфейса;
# - SSL-проверка отключена жёстко для обхода CERTIFICATE_VERIFY_FAILED;
# - токен берётся из .env/.evn рядом со скриптом.
#
# Установка:
#   pip install requests openpyxl certifi
#
# Файл .env рядом со скриптом:
#   TINVEST_TOKEN=твой_токен
#
# Запуск:
#   python jtrade.py

import os
import sys
import time
import uuid
import json
import math
import threading
import warnings
import ctypes
from pathlib import Path
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP, InvalidOperation

import requests
from openpyxl import Workbook, load_workbook

import tkinter as tk
from tkinter import ttk, messagebox
from tkinter.scrolledtext import ScrolledText

try:
    from urllib3.exceptions import InsecureRequestWarning
except Exception:
    InsecureRequestWarning = Warning


BASE_URL = "https://invest-public-api.tbank.ru/rest/tinkoff.public.invest.api.contract.v1"

APP_DIR = Path(__file__).resolve().parent
JOURNAL_PATH = APP_DIR / "trades_diary.xlsx"
STATE_PATH = APP_DIR / "active_trades.json"
ENV_FILE_USED: Path | None = None

# Dark palette close to the terminal screenshot.
BG = "#0F1A24"
BG_ALT = "#132231"
PANEL_BG = "#172838"
PANEL_BG_2 = "#1B2E3F"
HEADER_BG = "#0B141D"
INPUT_BG = "#223648"
INPUT_FG = "#F2EDE3"
FG = "#F2EDE3"
MUTED_FG = "#AEB8C2"
GREEN = "#3DDC97"
RED = "#FF5B6A"
YELLOW = "#F2C94C"
BLUE = "#5DADEC"
BORDER = "#31495F"
BUTTON_BG = "#102033"
BUTTON_ACTIVE = "#1E3A55"
BUY_BUTTON_BG = "#079B65"
BUY_BUTTON_ACTIVE = "#0BB978"
SELL_BUTTON_BG = "#A62A2A"
SELL_BUTTON_ACTIVE = "#C03535"
TREE_SELECTED = "#244A67"
FONT_FAMILY = "Calibri"

DEFAULT_RR = Decimal("2")
ENTRY_WAIT_SECONDS = 12
OCO_CHECK_SECONDS = 3

SSL_VERIFY = False
warnings.simplefilter("ignore", InsecureRequestWarning)


def load_local_env() -> None:
    global ENV_FILE_USED
    candidates = [APP_DIR / ".env", APP_DIR / ".evn"]
    env_path = next((p for p in candidates if p.exists()), None)
    if not env_path:
        return

    ENV_FILE_USED = env_path
    for raw_line in env_path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip()
        if not key:
            continue
        if len(value) >= 2 and value[0] == value[-1] and value[0] in {"'", '"'}:
            value = value[1:-1]
        os.environ[key] = value


load_local_env()
TOKEN = os.getenv("TINVEST_TOKEN") or os.getenv("INVEST_TOKEN")


def die(message: str):
    raise RuntimeError(message)


def headers() -> dict:
    if not TOKEN:
        die("Не задан TINVEST_TOKEN. Создай .env рядом со скриптом и добавь: TINVEST_TOKEN=твой_токен")
    return {
        "Authorization": f"Bearer {TOKEN}",
        "Content-Type": "application/json",
    }


def post(method: str, payload: dict) -> dict:
    url = f"{BASE_URL}.{method}"
    response = requests.post(url, headers=headers(), json=payload, timeout=25, verify=SSL_VERIFY)
    if response.status_code >= 400:
        raise RuntimeError(f"API {method} вернул {response.status_code}: {response.text}")
    return response.json()


def q_to_decimal(q: dict | None) -> Decimal:
    if not q:
        return Decimal("0")
    units = Decimal(str(q.get("units", "0") or "0"))
    nano = Decimal(str(q.get("nano", 0) or 0)) / Decimal("1000000000")
    return units + nano


def decimal_to_q(value: Decimal) -> dict:
    value = Decimal(value)
    sign = -1 if value < 0 else 1
    value_abs = abs(value)
    units = int(value_abs)
    nano = int((value_abs - Decimal(units)) * Decimal("1000000000"))
    if sign < 0:
        units = -units
        nano = -nano
    return {"units": str(units), "nano": nano}


def money_to_decimal(m: dict | None) -> Decimal:
    return q_to_decimal(m)


def round_to_step(price: Decimal, step: Decimal) -> Decimal:
    if step <= 0:
        return price
    return (price / step).quantize(Decimal("1"), rounding=ROUND_HALF_UP) * step


def parse_decimal(raw: str, field_name: str) -> Decimal:
    value = str(raw).strip().replace(",", ".")
    if not value:
        die(f"Поле «{field_name}» пустое.")
    try:
        result = Decimal(value)
    except (InvalidOperation, ValueError):
        die(f"Поле «{field_name}» должно быть числом. Пример: 110 или 109.85")
    if result <= 0:
        die(f"Поле «{field_name}» должно быть больше нуля.")
    return result


def fmt_dec(value, places: int = 2) -> str:
    try:
        d = Decimal(str(value))
        return f"{d.quantize(Decimal('1.' + '0' * places))}"
    except Exception:
        return str(value)


def fmt_money(value) -> str:
    try:
        d = Decimal(str(value))
        return f"{d.quantize(Decimal('1.00'))} ₽"
    except Exception:
        return str(value)


def fmt_percent(value) -> str:
    try:
        d = Decimal(str(value))
        return f"{d.quantize(Decimal('1.00'))}%"
    except Exception:
        return "—"


def signed_text(value: Decimal, suffix: str = " ₽") -> str:
    try:
        sign = "+" if value > 0 else ""
        return f"{sign}{fmt_dec(value)}{suffix}"
    except Exception:
        return "—"


def price_type_for(class_code: str) -> str:
    if str(class_code).upper() == "SPBFUT":
        return "PRICE_TYPE_POINT"
    return "PRICE_TYPE_CURRENCY"


def side_to_text(side: str) -> str:
    return "LONG" if side == "BUY" else "SHORT"


def position_side_from_qty(qty: Decimal) -> str:
    if qty > 0:
        return "BUY"
    if qty < 0:
        return "SELL"
    return "FLAT"


# ---------------------------- API helpers ----------------------------

def get_accounts() -> list[dict]:
    return post("UsersService/GetAccounts", {}).get("accounts", [])


def get_portfolio(account_id: str) -> dict:
    return post("OperationsService/GetPortfolio", {"accountId": account_id})


def get_positions(account_id: str) -> dict:
    return post("OperationsService/GetPositions", {"accountId": account_id})


def get_withdraw_limits(account_id: str) -> dict:
    return post("OperationsService/GetWithdrawLimits", {"accountId": account_id})


def get_stop_orders(account_id: str) -> list[dict]:
    return post("StopOrdersService/GetStopOrders", {"accountId": account_id}).get("stopOrders", [])


def find_instrument(ticker: str) -> dict:
    raw = ticker.strip().upper()
    preferred_class_code = None
    search_ticker = raw

    if "_" in raw:
        search_ticker, preferred_class_code = raw.split("_", 1)
        search_ticker = search_ticker.strip().upper()
        preferred_class_code = preferred_class_code.strip().upper()

    data = post("InstrumentsService/FindInstrument", {
        "query": search_ticker,
        "apiTradeAvailableFlag": True,
    })

    exact = []
    for inst in data.get("instruments", []):
        ticker_match = str(inst.get("ticker", "")).upper() == search_ticker
        class_match = not preferred_class_code or str(inst.get("classCode", "")).upper() == preferred_class_code
        if ticker_match and class_match:
            exact.append(inst)

    if not exact:
        die(
            f"Не нашёл торговый инструмент по тикеру {ticker}. "
            f"Для фьючерса можно вводить TICKER_CLASSCODE, например BRM6_SPBFUT."
        )

    exact.sort(key=lambda x: 0 if str(x.get("classCode", "")).upper() == "SPBFUT" else 1)
    return exact[0]


def get_instrument_id(inst: dict) -> str:
    instrument_id = inst.get("uid") or inst.get("instrumentUid") or inst.get("instrumentId") or inst.get("figi")
    if not instrument_id:
        die(f"У инструмента нет uid/figi. Ответ API: {inst}")
    return instrument_id


def get_instrument_full_by_uid(uid: str) -> dict:
    if not uid:
        return {}
    data = post("InstrumentsService/GetInstrumentBy", {
        "idType": "INSTRUMENT_ID_TYPE_UID",
        "id": uid,
    })
    return data.get("instrument", {})


def get_instrument_full_by_figi(figi: str) -> dict:
    if not figi:
        return {}
    data = post("InstrumentsService/GetInstrumentBy", {
        "idType": "INSTRUMENT_ID_TYPE_FIGI",
        "id": figi,
    })
    return data.get("instrument", {})


def get_min_step(instrument: dict, full: dict) -> Decimal:
    raw = full.get("minPriceIncrement") or instrument.get("minPriceIncrement")
    if raw:
        return q_to_decimal(raw)
    return Decimal("0.01")


def get_best_entry_price(instrument_id: str, side: str) -> Decimal:
    data = post("MarketDataService/GetOrderBook", {
        "instrumentId": instrument_id,
        "depth": 1,
    })
    bids = data.get("bids", [])
    asks = data.get("asks", [])

    if side == "BUY":
        if not asks:
            die("В стакане нет ask. Нельзя открыть LONG по текущей лимитной цене.")
        return q_to_decimal(asks[0]["price"])

    if side == "SELL":
        if not bids:
            die("В стакане нет bid. Нельзя открыть SHORT по текущей лимитной цене.")
        return q_to_decimal(bids[0]["price"])

    die("side должен быть BUY или SELL")


def calc_tp_sl(entry: Decimal, side: str, step: Decimal, sl_price: Decimal, tp_price: Decimal | None = None) -> tuple[Decimal, Decimal]:
    sl = round_to_step(sl_price, step)

    if side == "BUY":
        if sl >= entry:
            die(f"Для LONG стоп должен быть ниже входа. Вход: {entry}, SL: {sl}")
        if tp_price is None:
            tp = entry + ((entry - sl) * DEFAULT_RR)
        else:
            tp = tp_price
            if tp <= entry:
                die(f"Для LONG TP должен быть выше входа. Вход: {entry}, TP: {tp}")
    elif side == "SELL":
        if sl <= entry:
            die(f"Для SHORT стоп должен быть выше входа. Вход: {entry}, SL: {sl}")
        if tp_price is None:
            tp = entry - ((sl - entry) * DEFAULT_RR)
        else:
            tp = tp_price
            if tp >= entry:
                die(f"Для SHORT TP должен быть ниже входа. Вход: {entry}, TP: {tp}")
    else:
        die("side должен быть BUY или SELL")

    return round_to_step(tp, step), sl


def post_limit_entry(account_id: str, instrument_id: str, qty: int, price: Decimal, side: str, class_code: str) -> str:
    direction = "ORDER_DIRECTION_BUY" if side == "BUY" else "ORDER_DIRECTION_SELL"
    payload = {
        "accountId": account_id,
        "instrumentId": instrument_id,
        "quantity": str(qty),
        "price": decimal_to_q(price),
        "direction": direction,
        "orderType": "ORDER_TYPE_LIMIT",
        "orderId": str(uuid.uuid4()),
        "timeInForce": "TIME_IN_FORCE_FILL_AND_KILL",
        "priceType": price_type_for(class_code),
        "confirmMarginTrade": side == "SELL",
    }
    data = post("OrdersService/PostOrder", payload)
    order_id = data.get("orderId")
    if not order_id:
        die(f"Не получил orderId от API. Ответ: {data}")
    return order_id


def post_market_order(account_id: str, instrument_id: str, qty: int, direction: str, class_code: str, confirm_margin: bool = False) -> str:
    payload = {
        "accountId": account_id,
        "instrumentId": instrument_id,
        "quantity": str(qty),
        "price": decimal_to_q(Decimal("0")),
        "direction": direction,
        "orderType": "ORDER_TYPE_MARKET",
        "orderId": str(uuid.uuid4()),
        "timeInForce": "TIME_IN_FORCE_DAY",
        "priceType": price_type_for(class_code),
        "confirmMarginTrade": confirm_margin,
    }
    data = post("OrdersService/PostOrder", payload)
    order_id = data.get("orderId")
    if not order_id:
        die(f"Не получил orderId от API. Ответ: {data}")
    return order_id


def get_order_state(account_id: str, order_id: str) -> dict:
    return post("OrdersService/GetOrderState", {
        "accountId": account_id,
        "orderId": order_id,
    })


def wait_fill(account_id: str, order_id: str, timeout_sec: int = ENTRY_WAIT_SECONDS) -> int:
    last_state = None
    for _ in range(timeout_sec):
        state = get_order_state(account_id, order_id)
        last_state = state
        status = str(state.get("executionReportStatus", "")).upper()
        lots_executed = int(state.get("lotsExecuted", 0) or 0)
        if "FILL" in status and lots_executed > 0:
            return lots_executed
        if "REJECT" in status or "CANCEL" in status:
            return lots_executed
        time.sleep(1)
    return int((last_state or {}).get("lotsExecuted", 0) or 0)


def post_stop(account_id: str, instrument_id: str, qty: int, price: Decimal, original_side: str, stop_type: str, class_code: str) -> str:
    exit_direction = "STOP_ORDER_DIRECTION_SELL" if original_side == "BUY" else "STOP_ORDER_DIRECTION_BUY"
    payload = {
        "accountId": account_id,
        "instrumentId": instrument_id,
        "quantity": str(qty),
        "price": decimal_to_q(price),
        "stopPrice": decimal_to_q(price),
        "direction": exit_direction,
        "expirationType": "STOP_ORDER_EXPIRATION_TYPE_GOOD_TILL_CANCEL",
        "stopOrderType": stop_type,
        "exchangeOrderType": "EXCHANGE_ORDER_TYPE_MARKET",
        "priceType": price_type_for(class_code),
        "orderId": str(uuid.uuid4()),
        "confirmMarginTrade": False,
    }
    data = post("StopOrdersService/PostStopOrder", payload)
    stop_id = data.get("stopOrderId")
    if not stop_id:
        die(f"Не получил stopOrderId. Ответ: {data}")
    return stop_id


def cancel_stop_order(account_id: str, stop_order_id: str):
    if not stop_order_id:
        return
    post("StopOrdersService/CancelStopOrder", {
        "accountId": account_id,
        "stopOrderId": stop_order_id,
    })


# ---------------------------- State and journal ----------------------------

def load_state() -> dict:
    if not STATE_PATH.exists():
        return {"active_trades": []}
    try:
        return json.loads(STATE_PATH.read_text(encoding="utf-8"))
    except Exception:
        return {"active_trades": []}


def save_state(state: dict):
    STATE_PATH.write_text(json.dumps(state, ensure_ascii=False, indent=2), encoding="utf-8")


def add_active_trade(trade: dict):
    state = load_state()
    state.setdefault("active_trades", []).append(trade)
    save_state(state)


def update_active_trade(trade_id: str, updates: dict):
    state = load_state()
    for trade in state.get("active_trades", []):
        if trade.get("trade_id") == trade_id:
            trade.update(updates)
            break
    save_state(state)


def remove_active_trade(trade_id: str):
    state = load_state()
    state["active_trades"] = [t for t in state.get("active_trades", []) if t.get("trade_id") != trade_id]
    save_state(state)


def find_active_trade(account_id: str, ticker: str | None = None, side: str | None = None, instrument_id: str | None = None) -> dict | None:
    state = load_state()
    ticker_upper = ticker.upper().strip() if ticker else None
    for trade in reversed(state.get("active_trades", [])):
        if trade.get("account_id") != account_id:
            continue
        if instrument_id and trade.get("instrument_id") != instrument_id:
            continue
        if ticker_upper and str(trade.get("ticker", "")).upper() != ticker_upper:
            continue
        if side and trade.get("side") != side:
            continue
        return trade
    return None


def ensure_journal():
    headers_row = ["Дата", "Капитал счёта", "PnL", "Тикер", "Комментарии"]
    if JOURNAL_PATH.exists():
        wb = load_workbook(JOURNAL_PATH)
        ws = wb.active
        current = [cell.value for cell in ws[1]] if ws.max_row >= 1 else []
        if current[:5] != headers_row:
            ws = wb.create_sheet("Trades")
            ws.append(headers_row)
        wb.save(JOURNAL_PATH)
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Trades"
    ws.append(headers_row)
    widths = [14, 18, 14, 16, 60]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = width
    wb.save(JOURNAL_PATH)


def append_journal_row(date_str: str, capital: Decimal | None, pnl: Decimal | None, ticker: str, comments: str = ""):
    ensure_journal()
    wb = load_workbook(JOURNAL_PATH)
    ws = wb.active
    if [cell.value for cell in ws[1]][:5] != ["Дата", "Капитал счёта", "PnL", "Тикер", "Комментарии"]:
        ws = wb["Trades"] if "Trades" in wb.sheetnames else wb.create_sheet("Trades")
        if ws.max_row == 1 and not ws[1][0].value:
            ws.append(["Дата", "Капитал счёта", "PnL", "Тикер", "Комментарии"])

    ws.append([
        date_str,
        float(capital) if capital is not None else None,
        float(pnl) if pnl is not None else None,
        ticker,
        comments,
    ])
    wb.save(JOURNAL_PATH)


def get_total_portfolio_value(account_id: str) -> Decimal:
    return money_to_decimal(get_portfolio(account_id).get("totalAmountPortfolio"))


# ---------------------------- GUI styling ----------------------------

def set_windows_app_user_model_id() -> None:
    try:
        if sys.platform.startswith("win"):
            ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID("Kaiyah.JTrade.Dark.4")
    except Exception:
        pass


def find_app_icon() -> Path | None:
    candidates = [
        APP_DIR / "app_icon.ico",
        APP_DIR / "icon.ico",
        APP_DIR / "jeatrade.ico",
        APP_DIR / "app_icon.png",
        APP_DIR / "icon.png",
        APP_DIR / "jeatrade.png",
    ]
    return next((p for p in candidates if p.exists()), None)


def apply_window_icon(root: tk.Tk) -> None:
    icon_path = find_app_icon()
    if not icon_path:
        return
    try:
        if icon_path.suffix.lower() == ".ico":
            root.iconbitmap(default=str(icon_path))
        if icon_path.suffix.lower() in {".png", ".gif"}:
            img = tk.PhotoImage(file=str(icon_path))
            root.iconphoto(True, img)
            root._app_icon_photo = img
    except Exception:
        pass


def remove_titlebar_text(root: tk.Tk) -> None:
    try:
        root.title("")
    except Exception:
        pass


def apply_theme(root: tk.Tk) -> None:
    root.configure(bg=BG)
    root.option_add("*Font", f"{FONT_FAMILY} 10")

    style = ttk.Style(root)
    try:
        style.theme_use("clam")
    except tk.TclError:
        pass

    style.configure(".", font=(FONT_FAMILY, 10), background=BG, foreground=FG)
    style.configure("TFrame", background=BG)
    style.configure("Card.TFrame", background=PANEL_BG, relief="solid", borderwidth=1)
    style.configure("Header.TFrame", background=HEADER_BG)
    style.configure("Panel.TFrame", background=PANEL_BG)

    style.configure("TLabel", background=BG, foreground=FG, font=(FONT_FAMILY, 10))
    style.configure("Title.TLabel", background=BG, foreground=FG, font=(FONT_FAMILY, 20, "bold"))
    style.configure("CardTitle.TLabel", background=PANEL_BG, foreground=MUTED_FG, font=(FONT_FAMILY, 9, "bold"))
    style.configure("CardValue.TLabel", background=PANEL_BG, foreground=FG, font=(FONT_FAMILY, 15, "bold"))
    style.configure("Muted.TLabel", background=BG, foreground=MUTED_FG, font=(FONT_FAMILY, 9))
    style.configure("PanelMuted.TLabel", background=PANEL_BG, foreground=MUTED_FG, font=(FONT_FAMILY, 9))
    style.configure("Bold.TLabel", background=BG, foreground=FG, font=(FONT_FAMILY, 11, "bold"))
    style.configure("Green.TLabel", background=PANEL_BG, foreground=GREEN, font=(FONT_FAMILY, 15, "bold"))
    style.configure("Red.TLabel", background=PANEL_BG, foreground=RED, font=(FONT_FAMILY, 15, "bold"))

    style.configure(
        "TLabelframe",
        background=BG,
        foreground=FG,
        bordercolor=BORDER,
        lightcolor=BORDER,
        darkcolor=BORDER,
        relief="solid",
    )
    style.configure("TLabelframe.Label", background=BG, foreground=FG, font=(FONT_FAMILY, 11, "bold"))

    style.configure(
        "TEntry",
        fieldbackground=INPUT_BG,
        foreground=INPUT_FG,
        insertcolor=INPUT_FG,
        bordercolor=BORDER,
        lightcolor=BORDER,
        darkcolor=BORDER,
        padding=5,
    )
    style.configure(
        "TCombobox",
        fieldbackground=INPUT_BG,
        foreground=INPUT_FG,
        background=INPUT_BG,
        arrowcolor=INPUT_FG,
        bordercolor=BORDER,
        lightcolor=BORDER,
        darkcolor=BORDER,
        padding=5,
    )
    style.map("TCombobox", fieldbackground=[("readonly", INPUT_BG)], foreground=[("readonly", INPUT_FG)])

    style.configure(
        "TButton",
        background=BUTTON_BG,
        foreground=FG,
        font=(FONT_FAMILY, 10, "bold"),
        borderwidth=1,
        relief="solid",
        padding=(11, 7),
        bordercolor=BORDER,
        lightcolor=BORDER,
        darkcolor=BORDER,
    )
    style.map(
        "TButton",
        background=[("active", BUTTON_ACTIVE), ("pressed", BUTTON_ACTIVE)],
        foreground=[("active", FG), ("pressed", FG)],
    )

    style.configure(
        "Buy.TButton",
        background=BUY_BUTTON_BG,
        foreground="#FFFFFF",
        font=(FONT_FAMILY, 10, "bold"),
        borderwidth=0,
        relief="flat",
        padding=(11, 7),
    )
    style.map(
        "Buy.TButton",
        background=[("active", BUY_BUTTON_ACTIVE), ("pressed", BUY_BUTTON_ACTIVE)],
        foreground=[("active", "#FFFFFF"), ("pressed", "#FFFFFF")],
    )

    style.configure(
        "Sell.TButton",
        background=SELL_BUTTON_BG,
        foreground="#FFFFFF",
        font=(FONT_FAMILY, 10, "bold"),
        borderwidth=0,
        relief="flat",
        padding=(11, 7),
    )
    style.map(
        "Sell.TButton",
        background=[("active", SELL_BUTTON_ACTIVE), ("pressed", SELL_BUTTON_ACTIVE)],
        foreground=[("active", "#FFFFFF"), ("pressed", "#FFFFFF")],
    )

    style.configure("TNotebook", background=BG, borderwidth=0)
    style.configure("TNotebook.Tab", background=BUTTON_BG, foreground=MUTED_FG, padding=(16, 8), font=(FONT_FAMILY, 10, "bold"))
    style.map("TNotebook.Tab", background=[("selected", PANEL_BG_2)], foreground=[("selected", FG)])

    style.configure(
        "Treeview",
        background=PANEL_BG_2,
        foreground=FG,
        fieldbackground=PANEL_BG_2,
        rowheight=25,
        borderwidth=0,
        font=(FONT_FAMILY, 10),
    )
    style.configure(
        "Treeview.Heading",
        background=HEADER_BG,
        foreground=FG,
        relief="flat",
        font=(FONT_FAMILY, 10, "bold"),
    )
    style.map("Treeview", background=[("selected", TREE_SELECTED)], foreground=[("selected", FG)])
    style.map("Treeview.Heading", background=[("active", HEADER_BG)])

    style.configure("Vertical.TScrollbar", background=BUTTON_BG, troughcolor=BG, bordercolor=BORDER, arrowcolor=FG)
    style.configure("Horizontal.TScrollbar", background=BUTTON_BG, troughcolor=BG, bordercolor=BORDER, arrowcolor=FG)


def style_textbox(widget: ScrolledText) -> None:
    widget.configure(
        bg=PANEL_BG_2,
        fg=FG,
        insertbackground=FG,
        selectbackground=TREE_SELECTED,
        selectforeground=FG,
        font=(FONT_FAMILY, 9),
        relief="flat",
        borderwidth=0,
        padx=10,
        pady=8,
    )


def make_button(parent, text: str, command, width: int | None = None, style: str | None = None):
    btn = ttk.Button(parent, text=text, command=command, style=style or "TButton")
    if width:
        btn.configure(width=width)
    return btn


def make_tree(parent, columns: list[tuple[str, str, int, str]], height: int = 10) -> tuple[ttk.Treeview, ttk.Frame]:
    frame = ttk.Frame(parent)
    tree = ttk.Treeview(frame, columns=[c[0] for c in columns], show="headings", height=height)
    yscroll = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
    xscroll = ttk.Scrollbar(frame, orient="horizontal")

    separator_lines: list[tk.Frame] = []

    def redraw_column_separators(event=None):
        if not tree.winfo_exists():
            return
        if not separator_lines:
            return

        total_width = sum(int(tree.column(c[0], "width")) for c in columns)
        visible_width = max(1, tree.winfo_width())
        first_x = float(tree.xview()[0]) if tree.xview() else 0.0
        x_offset = int(first_x * total_width) if total_width > visible_width else 0

        x = 0
        height_px = max(1, tree.winfo_height())
        for idx, (key, _title, _width, _anchor) in enumerate(columns[:-1]):
            x += int(tree.column(key, "width"))
            visible_x = x - x_offset
            line = separator_lines[idx]
            if 0 < visible_x < visible_width:
                line.place(x=visible_x, y=0, width=1, height=height_px)
                line.lift()
            else:
                line.place_forget()

    def on_tree_xscroll(first, last):
        xscroll.set(first, last)
        tree.after_idle(redraw_column_separators)

    def on_horizontal_scroll(*args):
        tree.xview(*args)
        tree.after_idle(redraw_column_separators)

    xscroll.configure(command=on_horizontal_scroll)
    tree.configure(yscrollcommand=yscroll.set, xscrollcommand=on_tree_xscroll)

    for key, title, width, anchor in columns:
        tree.heading(key, text=title)
        tree.column(key, width=width, minwidth=60, anchor=anchor, stretch=True)

    for _ in range(max(0, len(columns) - 1)):
        separator_lines.append(tk.Frame(tree, bg=BORDER, width=1, height=1))

    tree.tag_configure("positive", foreground=GREEN)
    tree.tag_configure("negative", foreground=RED)
    tree.tag_configure("muted", foreground=MUTED_FG)
    tree.tag_configure("warning", foreground=YELLOW)

    tree.bind("<Configure>", redraw_column_separators, add="+")
    tree.bind("<ButtonRelease-1>", redraw_column_separators, add="+")
    tree.bind("<B1-Motion>", redraw_column_separators, add="+")
    tree.bind("<<TreeviewOpen>>", redraw_column_separators, add="+")
    tree.after(150, redraw_column_separators)

    tree.grid(row=0, column=0, sticky="nsew")
    yscroll.grid(row=0, column=1, sticky="ns")
    xscroll.grid(row=1, column=0, sticky="ew")
    frame.grid_rowconfigure(0, weight=1)
    frame.grid_columnconfigure(0, weight=1)
    return tree, frame


# ---------------------------- App ----------------------------

class JTradeDarkApp:
    def __init__(self, root: tk.Tk):
        self.root = root
        apply_theme(root)
        remove_titlebar_text(root)
        self.root.geometry("1380x800")
        self.root.minsize(1160, 680)

        self.accounts: list[dict] = []
        self.account_id: str | None = None
        self.account_label = ""
        self.selected_account_ids: list[str] = []
        self.account_labels_by_id: dict[str, str] = {}
        self.add_account_var = tk.StringVar()
        self.selected_accounts_var = tk.StringVar(value="Счета: —")

        self.oco_started = False
        self.instrument_cache: dict[str, dict] = {}
        self.portfolio_rows: dict[str, dict] = {}
        self.open_positions_rows: dict[str, dict] = {}
        self.active_trade_rows: dict[str, dict] = {}
        self.stop_order_rows: dict[str, dict] = {}

        self.total_var = tk.StringVar(value="—")
        self.expected_var = tk.StringVar(value="—")
        self.cash_var = tk.StringVar(value="—")
        self.blocked_var = tk.StringVar(value="—")
        self.var_margin_var = tk.StringVar(value="—")
        self.risk_var = tk.StringVar(value="—")
        self.account_var = tk.StringVar()

        ensure_journal()
        self.build_account_window()
        self.load_accounts_async()

    def run_async(self, target, on_done=None):
        def wrapper():
            try:
                result = target()
                if on_done:
                    self.root.after(0, lambda: on_done(result, None))
            except Exception as exc:
                if on_done:
                    self.root.after(0, lambda e=exc: on_done(None, e))
                else:
                    self.root.after(0, lambda e=exc: messagebox.showerror("Ошибка", str(e)))
        threading.Thread(target=wrapper, daemon=True).start()

    def log(self, text: str):
        if not hasattr(self, "log_box"):
            print(text)
            return
        self.log_box.configure(state="normal")
        self.log_box.insert("end", f"[{datetime.now().strftime('%H:%M:%S')}] {text}\n")
        self.log_box.see("end")
        self.log_box.configure(state="disabled")

    # ---------------------------- Account selection ----------------------------

    def account_short_label(self, account_id: str | None) -> str:
        if not account_id:
            return "—"
        if account_id in self.account_labels_by_id:
            return self.account_labels_by_id[account_id]
        for acc in self.accounts:
            if acc.get("id") == account_id:
                return self.make_account_label(acc, short=True)
        return str(account_id)

    def make_account_label(self, acc: dict, short: bool = False) -> str:
        name = acc.get("name") or "Счёт"
        account_id = acc.get("id") or ""
        if short:
            return f"{name} | {str(account_id)[-6:]}"
        return f"{name} | id={account_id} | type={acc.get('type')} | status={acc.get('status')} | access={acc.get('accessLevel')}"

    def get_selected_account_ids(self) -> list[str]:
        result = []
        for account_id in self.selected_account_ids:
            if account_id and account_id not in result:
                result.append(account_id)
        if self.account_id and self.account_id not in result:
            result.insert(0, self.account_id)
        return result

    def build_account_window(self):
        for widget in self.root.winfo_children():
            widget.destroy()

        outer = ttk.Frame(self.root, padding=24)
        outer.pack(fill="both", expand=True)

        card = ttk.Frame(outer, style="Card.TFrame", padding=18)
        card.pack(fill="x", padx=90, pady=(26, 16))

        ttk.Label(card, text="Брокерский счёт", style="CardTitle.TLabel").pack(anchor="w")
        self.account_combo = ttk.Combobox(card, textvariable=self.account_var, state="readonly", width=120)
        self.account_combo.pack(fill="x", pady=(8, 12))

        controls = ttk.Frame(card, style="Panel.TFrame")
        controls.pack(anchor="center")
        make_button(controls, "Обновить список", self.load_accounts_async).pack(side="left", padx=6)
        make_button(controls, "Открыть выбранный счёт", self.select_account).pack(side="left", padx=6)

        self.status_label = ttk.Label(outer, text="Получаю список брокерских счетов...", style="Muted.TLabel")
        self.status_label.pack(anchor="center", pady=(4, 8))

        log_frame = ttk.LabelFrame(outer, text="Лог", padding=10)
        log_frame.pack(fill="both", expand=True, padx=90)
        self.log_box = ScrolledText(log_frame, state="disabled", height=18)
        style_textbox(self.log_box)
        self.log_box.pack(fill="both", expand=True)

    def load_accounts_async(self):
        if hasattr(self, "status_label"):
            self.status_label.configure(text="Получаю список брокерских счетов...")
        self.log("Запрос UsersService/GetAccounts...")

        def done(result, error):
            if error:
                if hasattr(self, "status_label"):
                    self.status_label.configure(text="Ошибка получения счетов")
                self.log(str(error))
                messagebox.showerror("Ошибка", str(error))
                return

            self.accounts = result or []
            self.account_labels_by_id = {
                acc.get("id"): self.make_account_label(acc, short=True)
                for acc in self.accounts
                if acc.get("id")
            }
            if not self.accounts:
                self.status_label.configure(text="Счета не найдены")
                self.log("Счета не найдены.")
                return

            values = [f"{i}. {self.make_account_label(acc)}" for i, acc in enumerate(self.accounts, start=1)]
            if hasattr(self, "account_combo"):
                self.account_combo.configure(values=values)
                self.account_combo.current(0)
            if hasattr(self, "status_label"):
                self.status_label.configure(text=f"Найдено счетов: {len(values)}")
            self.log(f"Найдено счетов: {len(values)}")
            if hasattr(self, "additional_account_combo"):
                self.sync_account_header_controls()

        self.run_async(get_accounts, done)

    def select_account(self):
        idx = self.account_combo.current()
        if idx < 0 or idx >= len(self.accounts):
            messagebox.showwarning("Счёт", "Выбери счёт из списка.")
            return
        acc = self.accounts[idx]
        self.account_id = acc.get("id")
        self.account_label = self.make_account_label(acc, short=True)
        if not self.account_id:
            messagebox.showerror("Ошибка", "У выбранного счёта нет id.")
            return
        self.selected_account_ids = [self.account_id]
        self.build_main_window()
        self.refresh_all_async()
        self.start_oco_monitor()

    def sync_account_header_controls(self):
        selected_ids = self.get_selected_account_ids()
        labels = [self.account_short_label(account_id) for account_id in selected_ids]
        self.selected_accounts_var.set("Счета: " + "  +  ".join(labels))

        if not hasattr(self, "additional_account_combo"):
            return
        values = []
        self.additional_account_value_to_id = {}
        for acc in self.accounts:
            account_id = acc.get("id")
            if not account_id or account_id in selected_ids:
                continue
            label = self.make_account_label(acc, short=True)
            value = f"{label}"
            values.append(value)
            self.additional_account_value_to_id[value] = account_id
        self.additional_account_combo.configure(values=values)
        if values:
            self.additional_account_combo.set(values[0])
        else:
            self.additional_account_combo.set("")

    def add_second_account_async(self):
        if len(self.get_selected_account_ids()) >= 2:
            messagebox.showinfo("Счёт", "Сейчас подключается максимум два счёта: основной и второй.")
            return
        value = self.add_account_var.get().strip()
        account_id = getattr(self, "additional_account_value_to_id", {}).get(value)
        if not account_id:
            messagebox.showwarning("Счёт", "Выбери второй счёт из списка.")
            return
        if account_id not in self.selected_account_ids:
            self.selected_account_ids.append(account_id)
        self.sync_account_header_controls()
        self.log(f"Добавлен второй счёт: {self.account_short_label(account_id)}")
        self.refresh_all_async()

    def remove_second_account_async(self):
        if len(self.get_selected_account_ids()) <= 1:
            return
        removed = self.selected_account_ids.pop()
        self.sync_account_header_controls()
        self.log(f"Второй счёт убран из агрегации: {self.account_short_label(removed)}")
        self.refresh_all_async()

    # ---------------------------- Main window layout ----------------------------

    def build_main_window(self):
        for widget in self.root.winfo_children():
            widget.destroy()
        self.root.geometry("1380x800")

        root_frame = ttk.Frame(self.root, padding=12)
        root_frame.pack(fill="both", expand=True)

        header = ttk.Frame(root_frame, style="Header.TFrame", padding=(12, 10))
        header.pack(fill="x")
        header.grid_columnconfigure(0, weight=1)

        left = ttk.Frame(header, style="Header.TFrame")
        left.grid(row=0, column=0, sticky="ew")
        ttk.Label(left, textvariable=self.selected_accounts_var, style="Muted.TLabel").pack(side="left", padx=(0, 12))

        right = ttk.Frame(header, style="Header.TFrame")
        right.grid(row=0, column=1, sticky="e")
        ttk.Label(right, text="Добавить счёт:", style="Muted.TLabel").pack(side="left", padx=(0, 6))
        self.additional_account_combo = ttk.Combobox(right, textvariable=self.add_account_var, state="readonly", width=30)
        self.additional_account_combo.pack(side="left", padx=(0, 6))
        make_button(right, "Добавить", self.add_second_account_async).pack(side="left", padx=4)
        make_button(right, "Убрать второй", self.remove_second_account_async).pack(side="left", padx=4)
        make_button(right, "Обновить всё", self.refresh_all_async).pack(side="left", padx=4)
        make_button(right, "Сменить основной", self.build_account_window).pack(side="left", padx=4)
        self.sync_account_header_controls()

        cards = ttk.Frame(root_frame)
        cards.pack(fill="x", pady=(12, 10))
        self.make_card(cards, "Стоимость портфеля", self.total_var, 0)
        self.make_card(cards, "PnL портфеля", self.expected_var, 1)
        self.make_card(cards, "Свободные деньги", self.cash_var, 2)
        self.make_card(cards, "Заблокировано", self.blocked_var, 3)
        self.make_card(cards, "Ожид. вармаржа", self.var_margin_var, 4)
        self.make_card(cards, "Риск текущий, % капитала", self.risk_var, 5)

        self.notebook = ttk.Notebook(root_frame)
        self.notebook.pack(fill="both", expand=True)

        self.trade_tab = ttk.Frame(self.notebook, padding=10)
        self.portfolio_tab = ttk.Frame(self.notebook, padding=10)
        self.positions_tab = ttk.Frame(self.notebook, padding=10)
        self.stops_tab = ttk.Frame(self.notebook, padding=10)
        self.log_tab = ttk.Frame(self.notebook, padding=10)

        self.notebook.add(self.trade_tab, text="Сделка")
        self.notebook.add(self.portfolio_tab, text="Портфель")
        self.notebook.add(self.positions_tab, text="Открытые позиции")
        self.notebook.add(self.stops_tab, text="Заявки защиты")
        self.notebook.add(self.log_tab, text="Лог")

        self.build_trade_tab()
        self.build_portfolio_tab()
        self.build_positions_tab()
        self.build_stops_tab()
        self.build_log_tab()

        self.log(f"Основной счёт для новых сделок: {self.account_label}")
        self.log(f"Автожурнал: {JOURNAL_PATH}")

    def make_card(self, parent, label: str, variable: tk.StringVar, col: int):
        parent.grid_columnconfigure(col, weight=1, uniform="cards")
        card = ttk.Frame(parent, style="Card.TFrame", padding=12)
        card.grid(row=0, column=col, sticky="ew", padx=5)
        ttk.Label(card, text=label, style="CardTitle.TLabel").pack(anchor="w")
        ttk.Label(card, textvariable=variable, style="CardValue.TLabel").pack(anchor="w", pady=(4, 0))

    def build_trade_tab(self):
        layout = ttk.Frame(self.trade_tab)
        layout.pack(fill="both", expand=True)
        layout.grid_columnconfigure(0, weight=0)
        layout.grid_columnconfigure(1, weight=1)
        layout.grid_rowconfigure(1, weight=1)

        form = ttk.LabelFrame(layout, text="Открытие позиции", padding=14)
        form.grid(row=0, column=0, sticky="nsew", padx=(0, 10), pady=(0, 10))

        self.ticker_var = tk.StringVar()
        self.qty_var = tk.StringVar(value="1")
        self.sl_price_var = tk.StringVar()
        self.tp_manual_var = tk.StringVar()
        self.tp_preview_var = tk.StringVar(value="TP авто по RR 1:2")

        rows = [
            ("Тикер", self.ticker_var),
            ("Лоты/контракты", self.qty_var),
            ("SL цена", self.sl_price_var),
            ("TP цена вручную", self.tp_manual_var),
        ]
        for r, (label, var) in enumerate(rows):
            ttk.Label(form, text=label).grid(row=r, column=0, sticky="w", pady=5, padx=(0, 8))
            ttk.Entry(form, textvariable=var, width=22).grid(row=r, column=1, sticky="ew", pady=5)
        ttk.Label(form, textvariable=self.tp_preview_var, style="PanelMuted.TLabel").grid(row=4, column=0, columnspan=2, sticky="w", pady=(8, 8))

        make_button(form, "Купить", lambda: self.open_trade_async("BUY"), width=22, style="Buy.TButton").grid(row=5, column=0, columnspan=2, sticky="ew", pady=(10, 5))
        make_button(form, "Продать", lambda: self.open_trade_async("SELL"), width=22, style="Sell.TButton").grid(row=6, column=0, columnspan=2, sticky="ew", pady=5)
        make_button(form, "Обновить позицию по тикеру", self.refresh_current_ticker_position_async, width=22).grid(row=7, column=0, columnspan=2, sticky="ew", pady=(12, 5))

        self.current_ticker_status_var = tk.StringVar(value="Позиция по тикеру: —")
        status = ttk.LabelFrame(layout, text="Статус выбранного тикера", padding=12)
        status.grid(row=1, column=0, sticky="nsew", padx=(0, 10))
        ttk.Label(status, textvariable=self.current_ticker_status_var, wraplength=320, justify="left").pack(fill="x")

        quick = ttk.LabelFrame(layout, text="Позиции счёта", padding=10)
        quick.grid(row=0, column=1, rowspan=2, sticky="nsew")
        quick.grid_rowconfigure(0, weight=1)
        quick.grid_columnconfigure(0, weight=1)

        columns = [
            ("account", "Счёт", 120, "w"),
            ("ticker", "Тикер", 100, "w"),
            ("side", "Сторона", 80, "center"),
            ("qty", "Кол-во", 95, "e"),
            ("lots", "Лоты", 85, "e"),
            ("avg", "Средняя", 100, "e"),
            ("last", "Текущая", 100, "e"),
            ("pnl", "PnL", 105, "e"),
            ("pnl_pct", "PnL %", 80, "e"),
            ("var_margin", "Вармаржа", 105, "e"),
            ("risk_pct", "Риск %", 115, "e"),
            ("tp", "TP", 85, "e"),
            ("sl", "SL", 85, "e"),
        ]
        self.trade_positions_tree, frame = make_tree(quick, columns, height=18)
        frame.grid(row=0, column=0, sticky="nsew")
        self.trade_positions_tree.bind("<<TreeviewSelect>>", self.on_any_position_selected)

    def build_portfolio_tab(self):
        self.portfolio_tab.grid_columnconfigure(0, weight=1)
        self.portfolio_tab.grid_rowconfigure(0, weight=1)
        columns = [
            ("account", "Счёт", 120, "w"),
            ("section", "Раздел", 105, "w"),
            ("ticker", "Тикер", 100, "w"),
            ("name", "Название", 220, "w"),
            ("qty", "Кол-во", 100, "e"),
            ("lots", "Лоты", 85, "e"),
            ("avg", "Средняя", 105, "e"),
            ("last", "Текущая", 105, "e"),
            ("value", "Стоимость", 115, "e"),
            ("pnl", "Доход", 105, "e"),
            ("pnl_pct", "Доход %", 85, "e"),
            ("var_margin", "Вармаржа", 105, "e"),
            ("risk_pct", "Риск %", 115, "e"),
            ("figi", "FIGI/UID", 150, "w"),
        ]
        self.portfolio_tree, frame = make_tree(self.portfolio_tab, columns, height=21)
        frame.grid(row=0, column=0, sticky="nsew")

    def build_positions_tab(self):
        self.positions_tab.grid_columnconfigure(0, weight=1)
        self.positions_tab.grid_rowconfigure(0, weight=1)
        self.positions_tab.grid_rowconfigure(1, weight=0)

        columns = [
            ("account", "Счёт", 120, "w"),
            ("ticker", "Тикер", 100, "w"),
            ("side", "Сторона", 80, "center"),
            ("qty", "Кол-во", 95, "e"),
            ("lots", "Лоты", 85, "e"),
            ("avg", "Средняя", 105, "e"),
            ("last", "Текущая", 105, "e"),
            ("pnl", "PnL", 105, "e"),
            ("pnl_pct", "PnL %", 85, "e"),
            ("var_margin", "Вармаржа", 105, "e"),
            ("risk_pct", "Риск %", 115, "e"),
            ("tp", "TP", 85, "e"),
            ("sl", "SL", 85, "e"),
            ("state", "Состояние", 160, "w"),
        ]
        self.open_positions_tree, frame = make_tree(self.positions_tab, columns, height=19)
        frame.grid(row=0, column=0, sticky="nsew")
        self.open_positions_tree.bind("<<TreeviewSelect>>", self.on_any_position_selected)

        controls = ttk.LabelFrame(self.positions_tab, text="Управление выбранной позицией", padding=10)
        controls.grid(row=1, column=0, sticky="ew", pady=(10, 0))

        self.manage_selected_var = tk.StringVar(value="Выбрано: —")
        ttk.Label(controls, textvariable=self.manage_selected_var, style="Bold.TLabel").grid(row=0, column=0, columnspan=10, sticky="w", pady=(0, 8))

        ttk.Label(controls, text="Новый SL").grid(row=1, column=0, sticky="w", padx=(0, 6))
        self.manage_sl_var = tk.StringVar()
        ttk.Entry(controls, textvariable=self.manage_sl_var, width=12).grid(row=1, column=1, sticky="w", padx=(0, 12))

        ttk.Label(controls, text="Новый TP").grid(row=1, column=2, sticky="w", padx=(0, 6))
        self.manage_tp_var = tk.StringVar()
        ttk.Entry(controls, textvariable=self.manage_tp_var, width=12).grid(row=1, column=3, sticky="w", padx=(0, 12))

        make_button(controls, "Заменить TP/SL", self.replace_selected_protection_async).grid(row=1, column=4, sticky="ew", padx=4)
        make_button(controls, "Снять TP/SL", self.cancel_selected_protection_async).grid(row=1, column=5, sticky="ew", padx=4)
        make_button(controls, "Закрыть 25%", lambda: self.close_selected_position_async(Decimal("0.25"))).grid(row=1, column=6, sticky="ew", padx=4)
        make_button(controls, "Закрыть 50%", lambda: self.close_selected_position_async(Decimal("0.50"))).grid(row=1, column=7, sticky="ew", padx=4)
        make_button(controls, "Закрыть 100%", lambda: self.close_selected_position_async(Decimal("1"))).grid(row=1, column=8, sticky="ew", padx=4)
        make_button(controls, "Обновить всё", self.refresh_all_async).grid(row=1, column=9, sticky="ew", padx=4)

        for col in range(10):
            controls.grid_columnconfigure(col, weight=1 if col >= 4 else 0)

    def build_stops_tab(self):
        self.stops_tab.grid_columnconfigure(0, weight=1)
        self.stops_tab.grid_rowconfigure(0, weight=1)
        self.stops_tab.grid_rowconfigure(1, weight=1)

        active_frame = ttk.LabelFrame(self.stops_tab, text="Сделки, которыми управляет приложение", padding=8)
        active_frame.grid(row=0, column=0, sticky="nsew", pady=(0, 8))
        active_frame.grid_columnconfigure(0, weight=1)
        active_frame.grid_rowconfigure(0, weight=1)

        active_cols = [
            ("account", "Счёт", 120, "w"),
            ("created", "Создано", 145, "w"),
            ("ticker", "Тикер", 100, "w"),
            ("side", "Сторона", 80, "center"),
            ("qty", "Кол-во", 80, "e"),
            ("entry", "Вход", 100, "e"),
            ("tp", "TP", 100, "e"),
            ("sl", "SL", 100, "e"),
            ("tp_id", "TP id", 170, "w"),
            ("sl_id", "SL id", 170, "w"),
        ]
        self.active_trades_tree, frame1 = make_tree(active_frame, active_cols, height=8)
        frame1.grid(row=0, column=0, sticky="nsew")

        stops_frame = ttk.LabelFrame(self.stops_tab, text="Активные стоп-заявки у брокера", padding=8)
        stops_frame.grid(row=1, column=0, sticky="nsew")
        stops_frame.grid_columnconfigure(0, weight=1)
        stops_frame.grid_rowconfigure(0, weight=1)

        stop_cols = [
            ("account", "Счёт", 120, "w"),
            ("ticker", "Тикер", 110, "w"),
            ("type", "Тип", 170, "w"),
            ("direction", "Направление", 140, "w"),
            ("qty", "Кол-во", 90, "e"),
            ("price", "Цена", 100, "e"),
            ("stop", "Стоп цена", 100, "e"),
            ("id", "stopOrderId", 230, "w"),
        ]
        self.stop_orders_tree, frame2 = make_tree(stops_frame, stop_cols, height=9)
        frame2.grid(row=0, column=0, sticky="nsew")

    def build_log_tab(self):
        self.log_box = ScrolledText(self.log_tab, state="disabled")
        style_textbox(self.log_box)
        self.log_box.pack(fill="both", expand=True)

    # ---------------------------- Data refresh ----------------------------

    def refresh_all_async(self):
        account_ids = self.get_selected_account_ids()
        if not account_ids:
            return
        self.log("Обновляю агрегированный портфель, позиции и стоп-заявки...")

        def task():
            account_payloads = []
            all_rows = []
            all_stops = []
            total = Decimal("0")
            expected = Decimal("0")
            cash = Decimal("0")
            blocked = Decimal("0")
            var_margin = Decimal("0")

            # First pass: get all accounts and total selected capital.
            # Risk is calculated from current drawdown against this total capital,
            # not from SL distance and not from a single account only.
            for account_id in account_ids:
                portfolio = get_portfolio(account_id)
                positions = get_positions(account_id)
                try:
                    withdraw_limits = get_withdraw_limits(account_id)
                except Exception:
                    withdraw_limits = {}
                stops = self.enrich_stop_orders(get_stop_orders(account_id), account_id)
                account_total = money_to_decimal(portfolio.get("totalAmountPortfolio"))
                account_expected = money_to_decimal(portfolio.get("expectedYield"))
                account_cash, account_blocked = self.extract_cash(positions, withdraw_limits)
                account_var_margin = self.extract_var_margin(portfolio)
                total += account_total
                expected += account_expected
                cash += account_cash
                blocked += account_blocked
                var_margin += account_var_margin
                all_stops.extend(stops)
                account_payloads.append({
                    "account_id": account_id,
                    "portfolio": portfolio,
                    "stops": stops,
                    "account_total": account_total,
                    "account_label": self.account_short_label(account_id),
                })

            # Second pass: build rows after the aggregated capital is known.
            for payload in account_payloads:
                all_rows.extend(self.build_portfolio_rows(
                    payload["portfolio"],
                    payload["account_id"],
                    payload["account_label"],
                    payload["account_total"],
                    payload["stops"],
                    total,
                ))

            aggregated_rows = self.aggregate_position_rows(all_rows, total)
            risk_total = sum((row.get("risk_raw") or Decimal("0")) for row in all_rows)
            risk_pct = (risk_total / total * Decimal("100")) if total else Decimal("0")
            return {
                "rows": aggregated_rows,
                "raw_rows": all_rows,
                "stops": all_stops,
                "total": total,
                "expected": expected,
                "cash": cash,
                "blocked": blocked,
                "var_margin": var_margin,
                "risk_total": risk_total,
                "risk_pct": risk_pct,
            }

        def done(result, error):
            if error:
                self.log(f"Ошибка обновления: {error}")
                return
            self.render_summary(result["total"], result["expected"], result["cash"], result["blocked"], result["var_margin"], result["risk_pct"])
            self.render_portfolio(result["rows"])
            self.render_open_positions(result["rows"])
            self.render_active_trades()
            self.render_stop_orders(result["stops"])
            self.sync_account_header_controls()
            self.log("Данные обновлены.")

        self.run_async(task, done)

    def refresh_current_ticker_position_async(self):
        ticker = self.ticker_var.get().strip().upper()
        if not ticker:
            self.current_ticker_status_var.set("Позиция по тикеру: введи тикер")
            return

        account_ids = self.get_selected_account_ids()
        if not account_ids:
            return

        def task():
            account_payloads = []
            total = Decimal("0")
            for account_id in account_ids:
                portfolio = get_portfolio(account_id)
                stops = self.enrich_stop_orders(get_stop_orders(account_id), account_id)
                account_total = money_to_decimal(portfolio.get("totalAmountPortfolio"))
                total += account_total
                account_payloads.append((account_id, portfolio, stops, account_total))

            all_rows = []
            for account_id, portfolio, stops, account_total in account_payloads:
                all_rows.extend(self.build_portfolio_rows(
                    portfolio,
                    account_id,
                    self.account_short_label(account_id),
                    account_total,
                    stops,
                    total,
                ))
            rows = self.aggregate_position_rows(all_rows, total)
            for row in rows:
                if row["ticker"].upper() == ticker:
                    return row
            return None

        def done(row, error):
            if error:
                self.current_ticker_status_var.set(f"Ошибка: {error}")
                return
            if not row:
                self.current_ticker_status_var.set(f"Позиция {ticker}: нет открытой позиции")
                return
            self.current_ticker_status_var.set(
                f"Позиция {row['ticker']}: {row['side_text']} | счёт={row['account_label']} | "
                f"qty={row['qty']} | lots={row['qty_lots']} | avg={row['avg']} | "
                f"current={row['last']} | PnL={row['pnl']} ({row['pnl_pct']}) | риск={row['risk_pct']}"
            )

        self.run_async(task, done)

    def get_cached_instrument_for_position(self, position: dict) -> dict:
        uid = position.get("instrumentUid") or position.get("instrumentId") or position.get("positionUid") or ""
        figi = position.get("figi") or ""
        key = uid or figi
        if key and key in self.instrument_cache:
            return self.instrument_cache[key]

        inst = {}
        try:
            if uid:
                inst = get_instrument_full_by_uid(uid)
            elif figi:
                inst = get_instrument_full_by_figi(figi)
        except Exception:
            inst = {}
        if key:
            self.instrument_cache[key] = inst or {}
        return inst or {}

    def section_for_type(self, instrument_type: str) -> str:
        t = str(instrument_type).lower()
        if "currency" in t:
            return "Валюта"
        if "future" in t:
            return "Фьючерсы"
        if "option" in t:
            return "Опционы"
        if "bond" in t:
            return "Облигации"
        if "share" in t or "stock" in t:
            return "Акции"
        if "etf" in t:
            return "Фонды"
        return instrument_type or "Другое"

    def decimal_from_any_money(self, value) -> Decimal:
        if isinstance(value, dict):
            return money_to_decimal(value)
        if isinstance(value, (int, float, str, Decimal)):
            try:
                return Decimal(str(value))
            except Exception:
                return Decimal("0")
        return Decimal("0")

    def sum_money_items(self, items) -> Decimal:
        total = Decimal("0")
        if not items:
            return total
        if isinstance(items, dict):
            return self.decimal_from_any_money(items)
        if isinstance(items, list):
            for item in items:
                total += self.decimal_from_any_money(item)
        return total

    def extract_var_margin(self, portfolio: dict) -> Decimal:
        direct = self.decimal_from_any_money(portfolio.get("varMargin") or portfolio.get("var_margin"))
        if direct:
            return direct
        total = Decimal("0")
        for position in portfolio.get("positions", []):
            total += self.decimal_from_any_money(position.get("varMargin") or position.get("var_margin"))
        return total

    def stop_field(self, stop: dict, *names, default=None):
        for name in names:
            if name in stop and stop.get(name) not in (None, ""):
                return stop.get(name)
        return default

    def normalize_stop_order(self, stop: dict, account_id: str | None = None) -> dict:
        normalized = dict(stop)
        if account_id:
            normalized["_account_id"] = account_id
            normalized["_account_label"] = self.account_short_label(account_id)

        instrument_uid = self.stop_field(normalized, "instrumentUid", "instrumentId", "instrument_uid", "uid", default="")
        figi = self.stop_field(normalized, "figi", default="")
        ticker = str(self.stop_field(normalized, "ticker", default="") or "").upper()

        inst = {}
        cache_key = str(instrument_uid or figi or "")
        if cache_key and cache_key in self.instrument_cache:
            inst = self.instrument_cache.get(cache_key) or {}
        else:
            try:
                if instrument_uid:
                    inst = get_instrument_full_by_uid(str(instrument_uid))
                elif figi:
                    inst = get_instrument_full_by_figi(str(figi))
            except Exception:
                inst = {}
            if cache_key:
                self.instrument_cache[cache_key] = inst or {}

        if not ticker:
            ticker = str(inst.get("ticker") or "").upper()
        if not figi:
            figi = inst.get("figi") or ""
        if not instrument_uid:
            instrument_uid = inst.get("uid") or inst.get("instrumentUid") or ""

        normalized["_ticker"] = ticker
        normalized["_figi"] = str(figi or "")
        normalized["_instrument_uid"] = str(instrument_uid or "")
        normalized["_stop_id"] = self.stop_field(normalized, "stopOrderId", "stop_order_id", "orderId", "order_id", default="") or ""
        normalized["_stop_type"] = str(self.stop_field(normalized, "stopOrderType", "stop_order_type", default="") or "").upper()
        normalized["_direction"] = str(self.stop_field(normalized, "direction", default="") or "").upper()
        normalized["_stop_price"] = self.stop_price_value(normalized)
        return normalized

    def enrich_stop_orders(self, stops: list[dict], account_id: str) -> list[dict]:
        return [self.normalize_stop_order(stop, account_id) for stop in (stops or [])]

    def stop_price_value(self, stop: dict) -> Decimal:
        raw = self.stop_field(stop, "stopPrice", "stop_price", "activationPrice", "activation_price", "price")
        return q_to_decimal(raw) if isinstance(raw, dict) else self.decimal_from_any_money(raw)

    def stop_matches_position(self, stop: dict, row_ids: set[str], ticker: str) -> bool:
        stop_ids = {
            str(stop.get("_instrument_uid") or ""),
            str(stop.get("_figi") or ""),
            str(stop.get("instrumentUid") or ""),
            str(stop.get("instrumentId") or ""),
            str(stop.get("positionUid") or ""),
            str(stop.get("figi") or ""),
            str(stop.get("uid") or ""),
        }
        stop_ids = {x for x in stop_ids if x and x != "None"}
        clean_row_ids = {str(x) for x in row_ids if x and str(x) != "None"}
        if clean_row_ids and stop_ids and clean_row_ids.intersection(stop_ids):
            return True
        stop_ticker = str(stop.get("_ticker") or stop.get("ticker") or "").upper()
        return bool(stop_ticker and ticker and stop_ticker == ticker.upper())

    def classify_stop_kind(self, stop: dict, side: str, reference_price: Decimal) -> str | None:
        stop_type = str(stop.get("_stop_type") or stop.get("stopOrderType") or stop.get("stop_order_type") or "").upper()
        if "TAKE_PROFIT" in stop_type:
            return "tp"
        if "STOP_LOSS" in stop_type or "STOP_LIMIT" in stop_type or stop_type.endswith("STOP"):
            return "sl"

        # Fallback for broker responses that omit/rename stopOrderType:
        # classify by stop location relative to current/average price.
        price = stop.get("_stop_price") if isinstance(stop.get("_stop_price"), Decimal) else self.stop_price_value(stop)
        if price <= 0 or reference_price <= 0:
            return None
        if side == "BUY":
            return "tp" if price > reference_price else "sl"
        if side == "SELL":
            return "tp" if price < reference_price else "sl"
        return None

    def protection_from_broker_stops(self, account_id: str, ticker: str, side: str, row_ids: set[str], stops: list[dict], reference_price: Decimal) -> dict:
        result = {
            "tp_price": None,
            "sl_price": None,
            "tp_stop_id": "",
            "sl_stop_id": "",
            "source": "",
        }
        tp_candidates = []
        sl_candidates = []

        for raw_stop in stops or []:
            stop = raw_stop if raw_stop.get("_stop_price") is not None else self.normalize_stop_order(raw_stop, account_id)
            if stop.get("_account_id") and stop.get("_account_id") != account_id:
                continue
            if not self.stop_matches_position(stop, row_ids, ticker):
                continue
            price = stop.get("_stop_price") if isinstance(stop.get("_stop_price"), Decimal) else self.stop_price_value(stop)
            if price <= 0:
                continue
            stop_id = str(stop.get("_stop_id") or stop.get("stopOrderId") or stop.get("stop_order_id") or "")
            kind = self.classify_stop_kind(stop, side, reference_price)
            if kind == "tp":
                tp_candidates.append((price, stop_id))
            elif kind == "sl":
                sl_candidates.append((price, stop_id))

        if tp_candidates:
            # nearest TP to the current price is the one that matters first
            tp_candidates.sort(key=lambda item: abs(item[0] - reference_price))
            result["tp_price"], result["tp_stop_id"] = tp_candidates[0]
            result["source"] = "broker"
        if sl_candidates:
            # nearest SL to the current price is the one that matters first
            sl_candidates.sort(key=lambda item: abs(item[0] - reference_price))
            result["sl_price"], result["sl_stop_id"] = sl_candidates[0]
            result["source"] = "broker"
        return result

    def merge_protection(self, active_trade: dict | None, broker_protection: dict) -> dict:
        result = dict(broker_protection or {})
        if active_trade:
            try:
                if active_trade.get("tp_price") and active_trade.get("tp_price") != "—":
                    result["tp_price"] = Decimal(str(active_trade.get("tp_price")))
                    result["tp_stop_id"] = active_trade.get("tp_stop_id", "")
                    result["source"] = "app"
            except Exception:
                pass
            try:
                if active_trade.get("sl_price") and active_trade.get("sl_price") != "—":
                    result["sl_price"] = Decimal(str(active_trade.get("sl_price")))
                    result["sl_stop_id"] = active_trade.get("sl_stop_id", "")
                    result["source"] = "app"
            except Exception:
                pass
        return result

    def calc_current_drawdown_risk(self, pnl: Decimal, selected_capital_total: Decimal) -> tuple[Decimal, Decimal]:
        try:
            drawdown = -pnl if pnl < 0 else Decimal("0")
            risk_pct = (drawdown / selected_capital_total * Decimal("100")) if selected_capital_total else Decimal("0")
            return drawdown, risk_pct
        except Exception:
            return Decimal("0"), Decimal("0")

    def build_portfolio_rows(self, portfolio: dict, account_id: str, account_label: str, account_total: Decimal, stops: list[dict] | None = None, selected_capital_total: Decimal | None = None) -> list[dict]:
        rows = []
        for position in portfolio.get("positions", []):
            qty = q_to_decimal(position.get("quantity"))
            if qty == 0:
                continue

            inst = self.get_cached_instrument_for_position(position)
            figi = position.get("figi", "")
            instrument_uid = position.get("instrumentUid") or position.get("instrumentId") or inst.get("uid", "")
            position_uid = position.get("positionUid") or ""
            uid = instrument_uid or position_uid or ""
            ticker = inst.get("ticker") or figi or uid or "—"
            name = inst.get("name") or inst.get("title") or "—"
            class_code = inst.get("classCode") or position.get("classCode", "")
            instrument_type = position.get("instrumentType") or inst.get("instrumentType") or ""
            qty_lots = q_to_decimal(position.get("quantityLots")) if position.get("quantityLots") else qty
            avg = money_to_decimal(position.get("averagePositionPrice"))
            last = money_to_decimal(position.get("currentPrice"))
            pnl = money_to_decimal(position.get("expectedYield"))
            var_margin = self.decimal_from_any_money(position.get("varMargin") or position.get("var_margin"))
            value = last * qty
            base = abs(avg * qty) if avg and qty else Decimal("0")
            pnl_pct = (pnl / base * Decimal("100")) if base else Decimal("0")
            side = position_side_from_qty(qty)
            row_ids = {str(x) for x in (instrument_uid, position_uid, figi, uid) if x}
            active_trade = find_active_trade(account_id, ticker=ticker, side=None, instrument_id=uid) or find_active_trade(account_id, ticker=ticker, side=None, instrument_id=instrument_uid)
            reference_price = last if last > 0 else avg
            risk_capital = selected_capital_total if selected_capital_total is not None else account_total
            broker_protection = self.protection_from_broker_stops(account_id, ticker, side, row_ids, stops or [], reference_price)
            protection = self.merge_protection(active_trade, broker_protection)
            tp_price = protection.get("tp_price")
            sl_price = protection.get("sl_price")
            risk, risk_pct = self.calc_current_drawdown_risk(pnl, risk_capital)

            row = {
                "account_id": account_id,
                "account_label": account_label,
                "account_total_raw": account_total,
                "section": self.section_for_type(instrument_type),
                "instrument_type": instrument_type,
                "ticker": ticker,
                "name": name,
                "qty": qty,
                "qty_lots": qty_lots,
                "avg_raw": avg,
                "last_raw": last,
                "value_raw": value,
                "pnl_raw": pnl,
                "pnl_pct_raw": pnl_pct,
                "var_margin_raw": var_margin,
                "risk_raw": risk,
                "risk_pct_raw": risk_pct,
                "side": side,
                "side_text": side_to_text(side) if side in {"BUY", "SELL"} else "FLAT",
                "avg": fmt_dec(avg),
                "last": fmt_dec(last),
                "value": fmt_money(value),
                "pnl": signed_text(pnl),
                "pnl_pct": fmt_percent(pnl_pct),
                "var_margin": signed_text(var_margin),
                "risk": "",
                "risk_pct": fmt_percent(risk_pct),
                "figi": figi,
                "uid": uid,
                "instrument_uid": instrument_uid,
                "position_uid": position_uid,
                "instrument_id": uid or figi,
                "class_code": class_code,
                "tp": fmt_dec(tp_price) if isinstance(tp_price, Decimal) and tp_price > 0 else "—",
                "sl": fmt_dec(sl_price) if isinstance(sl_price, Decimal) and sl_price > 0 else "—",
                "tp_stop_id": protection.get("tp_stop_id", ""),
                "sl_stop_id": protection.get("sl_stop_id", ""),
                "protection_source": protection.get("source", ""),
                "active_trade": active_trade,
                "legs": [],
            }
            row["legs"] = [row]
            rows.append(row)
        return rows

    def aggregate_position_rows(self, rows: list[dict], selected_capital_total: Decimal | None = None) -> list[dict]:
        grouped: dict[tuple[str, str], list[dict]] = {}
        passthrough = []
        for row in rows:
            key = (row.get("instrument_id") or row.get("ticker"), row.get("side"))
            if row.get("side") in {"BUY", "SELL"}:
                grouped.setdefault(key, []).append(row)
            else:
                passthrough.append(row)

        result = []
        for legs in grouped.values():
            result.append(self.aggregate_leg_group(legs, selected_capital_total))
        result.extend(passthrough)
        result.sort(key=lambda r: (r.get("section", ""), r.get("ticker", ""), r.get("account_label", "")))
        return result

    def aggregate_leg_group(self, legs: list[dict], selected_capital_total: Decimal | None = None) -> dict:
        if len(legs) == 1:
            row = dict(legs[0])
            row["legs"] = legs
            return row

        first = legs[0]
        qty = sum((leg["qty"] for leg in legs), Decimal("0"))
        qty_lots = sum((leg["qty_lots"] for leg in legs), Decimal("0"))
        value = sum((leg["value_raw"] for leg in legs), Decimal("0"))
        pnl = sum((leg["pnl_raw"] for leg in legs), Decimal("0"))
        weight = sum((abs(leg["qty"]) for leg in legs), Decimal("0"))
        avg = (sum((leg["avg_raw"] * abs(leg["qty"]) for leg in legs), Decimal("0")) / weight) if weight else Decimal("0")
        last = (sum((leg["last_raw"] * abs(leg["qty"]) for leg in legs), Decimal("0")) / weight) if weight else Decimal("0")
        base = sum((abs(leg["avg_raw"] * leg["qty"]) for leg in legs), Decimal("0"))
        pnl_pct = (pnl / base * Decimal("100")) if base else Decimal("0")
        var_margin = sum(((leg.get("var_margin_raw") or Decimal("0")) for leg in legs), Decimal("0"))
        risk = sum(((leg.get("risk_raw") or Decimal("0")) for leg in legs), Decimal("0"))
        account_total = selected_capital_total if selected_capital_total is not None else sum({leg["account_id"]: leg.get("account_total_raw", Decimal("0")) for leg in legs}.values(), Decimal("0"))
        risk_pct = (risk / account_total * Decimal("100")) if account_total else Decimal("0")
        account_labels = []
        for leg in legs:
            if leg["account_label"] not in account_labels:
                account_labels.append(leg["account_label"])

        tp_values = {str(leg.get("tp", "—")) for leg in legs}
        sl_values = {str(leg.get("sl", "—")) for leg in legs}
        tp = next(iter(tp_values)) if len(tp_values) == 1 else "mixed"
        sl = next(iter(sl_values)) if len(sl_values) == 1 else "mixed"
        tp_stop_ids = [leg.get("tp_stop_id", "") for leg in legs if leg.get("tp_stop_id")]
        sl_stop_ids = [leg.get("sl_stop_id", "") for leg in legs if leg.get("sl_stop_id")]

        row = dict(first)
        row.update({
            "account_id": "AGGREGATED",
            "account_label": " + ".join(account_labels),
            "account_total_raw": account_total,
            "qty": qty,
            "qty_lots": qty_lots,
            "avg_raw": avg,
            "last_raw": last,
            "value_raw": value,
            "pnl_raw": pnl,
            "pnl_pct_raw": pnl_pct,
            "var_margin_raw": var_margin,
            "risk_raw": risk if risk else None,
            "risk_pct_raw": risk_pct,
            "avg": fmt_dec(avg),
            "last": fmt_dec(last),
            "value": fmt_money(value),
            "pnl": signed_text(pnl),
            "pnl_pct": fmt_percent(pnl_pct),
            "var_margin": signed_text(var_margin),
            "risk": "",
            "risk_pct": fmt_percent(risk_pct),
            "tp": tp,
            "sl": sl,
            "tp_stop_id": ",".join(tp_stop_ids),
            "sl_stop_id": ",".join(sl_stop_ids),
            "protection_source": "mixed" if tp_stop_ids or sl_stop_ids else "",
            "active_trade": None,
            "legs": legs,
        })
        return row

    def extract_cash(self, positions: dict, withdraw_limits: dict | None = None) -> tuple[Decimal, Decimal]:
        withdraw_limits = withdraw_limits or {}
        cash = self.sum_money_items(positions.get("money"))
        blocked = Decimal("0")
        blocked += self.sum_money_items(positions.get("blocked"))
        blocked += self.sum_money_items(withdraw_limits.get("blocked"))
        blocked += self.sum_money_items(withdraw_limits.get("blockedGuarantee"))
        blocked += self.sum_money_items(withdraw_limits.get("blocked_guarantee"))
        return cash, blocked

    def render_summary(self, total: Decimal, expected: Decimal, cash: Decimal, blocked: Decimal, var_margin: Decimal, risk_pct: Decimal):
        self.total_var.set(fmt_money(total))
        self.expected_var.set(signed_text(expected))
        self.cash_var.set(fmt_money(cash))
        self.blocked_var.set(fmt_money(blocked))
        self.var_margin_var.set(signed_text(var_margin))
        self.risk_var.set(fmt_percent(risk_pct))

    def clear_tree(self, tree: ttk.Treeview):
        for item in tree.get_children():
            tree.delete(item)

    def row_state_text(self, row: dict) -> str:
        legs = row.get("legs") or [row]
        protected = 0
        for leg in legs:
            has_ids = bool(leg.get("tp_stop_id") or leg.get("sl_stop_id"))
            trade = leg.get("active_trade") or find_active_trade(leg["account_id"], ticker=leg["ticker"], side=leg["side"], instrument_id=leg["instrument_id"])
            if has_ids or (trade and (trade.get("tp_stop_id") or trade.get("sl_stop_id"))):
                protected += 1
        if protected == len(legs):
            return "TP/SL есть"
        if protected == 0:
            return "без TP/SL"
        return "частично защищено"

    def render_portfolio(self, rows: list[dict]):
        self.portfolio_rows.clear()
        self.clear_tree(self.portfolio_tree)
        for idx, row in enumerate(rows):
            iid = f"portfolio_{idx}"
            tag = "positive" if row["pnl_raw"] > 0 else "negative" if row["pnl_raw"] < 0 else "muted"
            self.portfolio_rows[iid] = row
            self.portfolio_tree.insert("", "end", iid=iid, values=(
                row["account_label"], row["section"], row["ticker"], row["name"], fmt_dec(row["qty"], 4), fmt_dec(row["qty_lots"], 4),
                row["avg"], row["last"], row["value"], row["pnl"], row["pnl_pct"], row["var_margin"], row["risk_pct"], row["figi"] or row["uid"],
            ), tags=(tag,))

    def render_open_positions(self, rows: list[dict]):
        self.open_positions_rows.clear()
        self.clear_tree(self.open_positions_tree)
        self.clear_tree(self.trade_positions_tree)

        open_rows = [r for r in rows if r["side"] in {"BUY", "SELL"} and str(r["instrument_type"]).lower() != "currency"]
        for idx, row in enumerate(open_rows):
            iid = f"open_{idx}"
            state = self.row_state_text(row)
            tag = "positive" if row["pnl_raw"] > 0 else "negative" if row["pnl_raw"] < 0 else "muted"
            values = (
                row["account_label"], row["ticker"], row["side_text"], fmt_dec(row["qty"], 4), fmt_dec(row["qty_lots"], 4),
                row["avg"], row["last"], row["pnl"], row["pnl_pct"], row["var_margin"], row["risk_pct"], row["tp"], row["sl"], state,
            )
            self.open_positions_rows[iid] = row
            self.open_positions_tree.insert("", "end", iid=iid, values=values, tags=(tag,))
            self.trade_positions_tree.insert("", "end", iid=iid, values=values[:-1], tags=(tag,))

    def render_active_trades(self):
        self.active_trade_rows.clear()
        self.clear_tree(self.active_trades_tree)
        selected_ids = set(self.get_selected_account_ids())
        state = load_state()
        for idx, trade in enumerate(state.get("active_trades", [])):
            if trade.get("account_id") not in selected_ids:
                continue
            iid = f"active_{idx}"
            self.active_trade_rows[iid] = trade
            self.active_trades_tree.insert("", "end", iid=iid, values=(
                self.account_short_label(trade.get("account_id")),
                trade.get("created_at", "—"),
                trade.get("ticker", "—"),
                side_to_text(trade.get("side", "BUY")),
                trade.get("qty", "—"),
                trade.get("entry_price", "—"),
                trade.get("tp_price", "—"),
                trade.get("sl_price", "—"),
                trade.get("tp_stop_id", "—"),
                trade.get("sl_stop_id", "—"),
            ))

    def render_stop_orders(self, stops: list[dict]):
        self.stop_order_rows.clear()
        self.clear_tree(self.stop_orders_tree)
        for idx, stop in enumerate(stops):
            iid = f"stop_{idx}"
            self.stop_order_rows[iid] = stop
            self.stop_orders_tree.insert("", "end", iid=iid, values=(
                stop.get("_account_label", "—"),
                stop.get("_ticker") or stop.get("ticker") or stop.get("figi") or stop.get("instrumentUid") or "—",
                stop.get("_stop_type") or stop.get("stopOrderType", "—"),
                stop.get("_direction") or stop.get("direction", "—"),
                stop.get("lotsRequested") or stop.get("quantity") or "—",
                fmt_dec(q_to_decimal(stop.get("price"))),
                fmt_dec(stop.get("_stop_price") if isinstance(stop.get("_stop_price"), Decimal) else self.stop_price_value(stop)),
                stop.get("_stop_id") or stop.get("stopOrderId", "—"),
            ))

    # ---------------------------- Selection and management ----------------------------

    def on_any_position_selected(self, event=None):
        row = self.get_selected_position_row()
        if not row:
            self.manage_selected_var.set("Выбрано: —")
            return
        self.manage_selected_var.set(
            f"Выбрано: {row['ticker']} {row['side_text']} | счёт={row['account_label']} | "
            f"lots={fmt_dec(row['qty_lots'], 4)} | avg={row['avg']} | PnL={row['pnl']} | риск={row['risk_pct']}"
        )
        if row.get("sl") and row["sl"] not in {"—", "mixed"}:
            self.manage_sl_var.set(str(row["sl"]))
        if row.get("tp") and row["tp"] not in {"—", "mixed"}:
            self.manage_tp_var.set(str(row["tp"]))

    def get_selected_position_row(self) -> dict | None:
        if hasattr(self, "open_positions_tree"):
            selected = self.open_positions_tree.selection()
            if selected and selected[0] in self.open_positions_rows:
                return self.open_positions_rows[selected[0]]
        if hasattr(self, "trade_positions_tree"):
            selected = self.trade_positions_tree.selection()
            if selected and selected[0] in self.open_positions_rows:
                return self.open_positions_rows[selected[0]]
        return None

    def get_ticker_qty_inputs(self) -> tuple[str, int]:
        ticker = self.ticker_var.get().strip().upper()
        if not ticker:
            die("Введи тикер.")
        try:
            qty = int(self.qty_var.get().strip())
        except ValueError:
            die("Количество должно быть целым числом.")
        if qty <= 0:
            die("Количество должно быть больше нуля.")
        return ticker, qty

    def parse_optional_tp(self, value: str) -> Decimal | None:
        raw = str(value).strip().replace(",", ".")
        if not raw:
            return None
        return parse_decimal(raw, "TP цена")

    def open_trade_async(self, side: str):
        if not self.account_id:
            messagebox.showwarning("Счёт", "Сначала выбери основной счёт.")
            return
        try:
            ticker, qty = self.get_ticker_qty_inputs()
            sl_price = parse_decimal(self.sl_price_var.get(), "SL цена")
            tp_manual = self.parse_optional_tp(self.tp_manual_var.get())
        except Exception as exc:
            messagebox.showerror("Ошибка", str(exc))
            return

        side_name = side_to_text(side)
        tp_text = str(tp_manual) if tp_manual is not None else f"авто RR 1:{DEFAULT_RR.normalize()}"
        if not messagebox.askyesno(
            "Подтверждение",
            f"Открыть {side_name} {ticker}, количество {qty}?\n"
            f"Счёт: {self.account_label}\nSL: {sl_price}\nTP: {tp_text}",
        ):
            return

        self.log(f"Готовлю открытие {side_name}: {ticker}, qty={qty}, account={self.account_label}, SL={sl_price}, TP={tp_text}")

        def task():
            before_portfolio = get_total_portfolio_value(self.account_id)
            inst = find_instrument(ticker)
            instrument_id = get_instrument_id(inst)
            class_code = inst.get("classCode", "")
            full = get_instrument_full_by_uid(instrument_id)
            step = get_min_step(inst, full)
            entry = round_to_step(get_best_entry_price(instrument_id, side), step)
            tp, sl = calc_tp_sl(entry, side, step, sl_price, tp_manual)

            order_id = post_limit_entry(self.account_id, instrument_id, qty, entry, side, class_code)
            filled_qty = wait_fill(self.account_id, order_id, ENTRY_WAIT_SECONDS)
            if filled_qty <= 0:
                return {"status": "not_filled", "ticker": ticker, "entry": entry, "order_id": order_id}

            tp_id = post_stop(self.account_id, instrument_id, filled_qty, tp, side, "STOP_ORDER_TYPE_TAKE_PROFIT", class_code)
            sl_id = post_stop(self.account_id, instrument_id, filled_qty, sl, side, "STOP_ORDER_TYPE_STOP_LOSS", class_code)

            trade = {
                "trade_id": str(uuid.uuid4()),
                "created_at": datetime.now().isoformat(timespec="seconds"),
                "account_id": self.account_id,
                "ticker": inst.get("ticker", ticker),
                "class_code": class_code,
                "instrument_id": instrument_id,
                "side": side,
                "qty": filled_qty,
                "entry_price": str(entry),
                "tp_price": str(tp),
                "sl_price": str(sl),
                "rr": str(DEFAULT_RR),
                "tp_stop_id": tp_id,
                "sl_stop_id": sl_id,
                "entry_portfolio": str(before_portfolio),
                "managed_external": False,
            }
            add_active_trade(trade)
            return {
                "status": "ok",
                "ticker": inst.get("ticker", ticker),
                "side_name": side_name,
                "requested_qty": qty,
                "qty": filled_qty,
                "entry": entry,
                "tp": tp,
                "sl": sl,
            }

        def done(result, error):
            if error:
                self.log(f"Ошибка открытия: {error}")
                messagebox.showerror("Ошибка открытия", str(error))
                return
            if result["status"] == "not_filled":
                self.log(f"Вход не исполнился: {result['ticker']} по {result['entry']}. TP/SL не ставились.")
                messagebox.showinfo("Не исполнено", "Входная лимитка не исполнилась. TP/SL не ставились.")
                self.refresh_all_async()
                return
            self.log(
                f"Открыт {result['side_name']} {result['ticker']} qty={result['qty']} "
                f"entry={result['entry']} TP={result['tp']} SL={result['sl']}"
            )
            if result["qty"] < result["requested_qty"]:
                self.log(f"Частичное исполнение: {result['qty']} из {result['requested_qty']}.")
            messagebox.showinfo("Готово", f"Открыто: {result['side_name']} {result['ticker']}\nTP: {result['tp']}\nSL: {result['sl']}")
            self.refresh_all_async()

        self.run_async(task, done)

    def row_legs(self, row: dict) -> list[dict]:
        legs = row.get("legs") or [row]
        return [leg for leg in legs if leg.get("account_id") and leg.get("account_id") != "AGGREGATED"]

    def leg_stop_ids(self, leg: dict, trade: dict | None = None) -> list[str]:
        ids = []
        for raw in (leg.get("tp_stop_id"), leg.get("sl_stop_id")):
            for stop_id in str(raw or "").split(","):
                stop_id = stop_id.strip()
                if stop_id and stop_id not in ids:
                    ids.append(stop_id)
        if trade:
            for raw in (trade.get("tp_stop_id"), trade.get("sl_stop_id")):
                for stop_id in str(raw or "").split(","):
                    stop_id = stop_id.strip()
                    if stop_id and stop_id not in ids:
                        ids.append(stop_id)
        return ids

    def selected_close_lots(self, row: dict, fraction: Decimal) -> int:
        lots = abs(row.get("qty_lots", Decimal("0")))
        if lots <= 0:
            die("Не вижу лоты по выбранной позиции.")
        qty = int(math.floor(float(lots * fraction)))
        return max(1, min(int(lots), qty))

    def close_selected_position_async(self, fraction: Decimal):
        row = self.get_selected_position_row()
        if not row:
            messagebox.showwarning("Позиция", "Выбери позицию в таблице открытых позиций.")
            return
        legs = self.row_legs(row)
        try:
            plan = [(leg, self.selected_close_lots(leg, fraction)) for leg in legs]
        except Exception as exc:
            messagebox.showerror("Ошибка", str(exc))
            return

        total_lots = sum(qty for _leg, qty in plan)
        side = row["side"]
        close_direction = "ORDER_DIRECTION_SELL" if side == "BUY" else "ORDER_DIRECTION_BUY"
        percent_text = fmt_percent(fraction * Decimal("100"))
        if not messagebox.askyesno(
            "Подтверждение",
            f"Закрыть {percent_text} позиции {row['ticker']} {row['side_text']} по рынку?\n"
            f"Счёт/счета: {row['account_label']}\nВсего к закрытию: {total_lots} лот(ов).",
        ):
            return

        self.log(f"Закрываю {row['ticker']} {row['side_text']} по рынку: accounts={row['account_label']}, lots={total_lots}")

        def task():
            results = []
            for leg, close_qty in plan:
                account_id = leg["account_id"]
                trade = find_active_trade(account_id, ticker=leg["ticker"], side=leg["side"], instrument_id=leg["instrument_id"])
                if fraction >= Decimal("1"):
                    for stop_id in self.leg_stop_ids(leg, trade):
                        try:
                            cancel_stop_order(account_id, stop_id)
                        except Exception:
                            pass

                order_id = post_market_order(
                    account_id,
                    leg["instrument_id"],
                    close_qty,
                    close_direction,
                    leg.get("class_code", ""),
                    confirm_margin=False,
                )
                filled = wait_fill(account_id, order_id, timeout_sec=20)
                after_portfolio = get_total_portfolio_value(account_id)
                pnl = None
                if trade and trade.get("entry_portfolio"):
                    pnl = after_portfolio - Decimal(str(trade["entry_portfolio"]))
                    if fraction >= Decimal("1"):
                        remove_active_trade(trade["trade_id"])

                append_journal_row(
                    datetime.now().strftime("%d.%m.%Y"),
                    after_portfolio,
                    pnl,
                    leg["ticker"],
                    f"Manual market close {percent_text}; account={self.account_short_label(account_id)}; filled={filled}; order={order_id}",
                )
                results.append((account_id, filled, after_portfolio, pnl))
            return results

        def done(result, error):
            if error:
                self.log(f"Ошибка закрытия: {error}")
                messagebox.showerror("Ошибка закрытия", str(error))
                return
            filled_total = sum(item[1] for item in result)
            self.log(f"Закрытие {row['ticker']}: исполнено суммарно={filled_total} лот(ов).")
            self.refresh_all_async()

        self.run_async(task, done)

    def cancel_selected_protection_async(self):
        row = self.get_selected_position_row()
        if not row:
            messagebox.showwarning("Позиция", "Выбери позицию в таблице открытых позиций.")
            return
        legs = self.row_legs(row)

        cancel_plan = []
        for leg in legs:
            trade = find_active_trade(leg["account_id"], ticker=leg["ticker"], side=leg["side"], instrument_id=leg["instrument_id"])
            stop_ids = self.leg_stop_ids(leg, trade)
            if stop_ids:
                cancel_plan.append((leg, trade, stop_ids))

        if not cancel_plan:
            messagebox.showinfo("TP/SL", "Для выбранной позиции нет активных TP/SL.")
            return
        if not messagebox.askyesno("Подтверждение", f"Снять TP/SL по {row['ticker']} на {len(cancel_plan)} счёте/ноге?"):
            return

        def task():
            for leg, trade, stop_ids in cancel_plan:
                for stop_id in stop_ids:
                    try:
                        cancel_stop_order(leg["account_id"], stop_id)
                    except Exception:
                        pass
                if trade:
                    update_active_trade(trade["trade_id"], {"tp_stop_id": "", "sl_stop_id": "", "tp_price": "—", "sl_price": "—"})
            return True

        def done(result, error):
            if error:
                self.log(f"Ошибка снятия TP/SL: {error}")
                messagebox.showerror("Ошибка", str(error))
                return
            self.log(f"TP/SL по {row['ticker']} сняты.")
            self.refresh_all_async()

        self.run_async(task, done)

    def replace_selected_protection_async(self):
        row = self.get_selected_position_row()
        if not row:
            messagebox.showwarning("Позиция", "Выбери позицию в таблице открытых позиций.")
            return
        try:
            sl = parse_decimal(self.manage_sl_var.get(), "Новый SL")
            tp_manual = self.parse_optional_tp(self.manage_tp_var.get())
        except Exception as exc:
            messagebox.showerror("Ошибка", str(exc))
            return

        legs = self.row_legs(row)
        tp_text = str(tp_manual) if tp_manual is not None else f"авто RR 1:{DEFAULT_RR.normalize()} от средней каждой ноги"
        if not messagebox.askyesno(
            "Подтверждение",
            f"Заменить защиту по {row['ticker']}?\nСчёт/счета: {row['account_label']}\nSL: {sl}\nTP: {tp_text}",
        ):
            return

        def task():
            updates = []
            for leg in legs:
                side = leg["side"]
                entry = leg["avg_raw"] if leg["avg_raw"] > 0 else leg["last_raw"]
                if entry <= 0:
                    die(f"Не вижу среднюю/текущую цену позиции {leg['ticker']} на счёте {self.account_short_label(leg['account_id'])}.")
                if leg.get("uid"):
                    full = get_instrument_full_by_uid(leg["uid"])
                elif leg.get("figi"):
                    full = get_instrument_full_by_figi(leg["figi"])
                else:
                    full = {}
                step = get_min_step({}, full)
                tp, sl_rounded = calc_tp_sl(entry, side, step, sl, tp_manual)
                qty = int(abs(leg["qty_lots"]))
                if qty <= 0:
                    die("Не вижу количество лотов для защитной заявки.")

                old_trade = find_active_trade(leg["account_id"], ticker=leg["ticker"], side=side, instrument_id=leg["instrument_id"])
                for stop_id in self.leg_stop_ids(leg, old_trade):
                    try:
                        cancel_stop_order(leg["account_id"], stop_id)
                    except Exception:
                        pass

                tp_id = post_stop(leg["account_id"], leg["instrument_id"], qty, tp, side, "STOP_ORDER_TYPE_TAKE_PROFIT", leg.get("class_code", ""))
                sl_id = post_stop(leg["account_id"], leg["instrument_id"], qty, sl_rounded, side, "STOP_ORDER_TYPE_STOP_LOSS", leg.get("class_code", ""))

                if old_trade:
                    update_active_trade(old_trade["trade_id"], {
                        "qty": qty,
                        "entry_price": str(entry),
                        "tp_price": str(tp),
                        "sl_price": str(sl_rounded),
                        "tp_stop_id": tp_id,
                        "sl_stop_id": sl_id,
                    })
                else:
                    before_portfolio = get_total_portfolio_value(leg["account_id"])
                    add_active_trade({
                        "trade_id": str(uuid.uuid4()),
                        "created_at": datetime.now().isoformat(timespec="seconds"),
                        "account_id": leg["account_id"],
                        "ticker": leg["ticker"],
                        "class_code": leg.get("class_code", ""),
                        "instrument_id": leg["instrument_id"],
                        "side": side,
                        "qty": qty,
                        "entry_price": str(entry),
                        "tp_price": str(tp),
                        "sl_price": str(sl_rounded),
                        "rr": str(DEFAULT_RR),
                        "tp_stop_id": tp_id,
                        "sl_stop_id": sl_id,
                        "entry_portfolio": str(before_portfolio),
                        "managed_external": True,
                    })
                updates.append((leg["account_id"], tp, sl_rounded))
            return updates

        def done(result, error):
            if error:
                self.log(f"Ошибка замены TP/SL: {error}")
                messagebox.showerror("Ошибка TP/SL", str(error))
                return
            details = "; ".join(f"{self.account_short_label(account_id)} TP={tp} SL={sl}" for account_id, tp, sl in result)
            self.log(f"Защита по {row['ticker']} заменена: {details}")
            self.refresh_all_async()

        self.run_async(task, done)

    # ---------------------------- OCO monitor ----------------------------

    def start_oco_monitor(self):
        if self.oco_started:
            return
        self.oco_started = True

        def loop():
            while True:
                try:
                    for account_id in self.get_selected_account_ids():
                        self.check_oco_once_for_account(account_id)
                except Exception as exc:
                    self.root.after(0, lambda e=exc: self.log(f"OCO-monitor: {e}"))
                time.sleep(OCO_CHECK_SECONDS)

        threading.Thread(target=loop, daemon=True).start()

    def check_oco_once_for_account(self, account_id: str):
        state = load_state()
        trades = state.get("active_trades", [])
        if not trades or not account_id:
            return

        active_stops = get_stop_orders(account_id)
        active_ids = {x.get("stopOrderId") for x in active_stops}

        changed = False
        for trade in list(trades):
            if trade.get("account_id") != account_id:
                continue

            tp_id = trade.get("tp_stop_id")
            sl_id = trade.get("sl_stop_id")
            if not tp_id and not sl_id:
                continue

            tp_active = tp_id in active_ids if tp_id else False
            sl_active = sl_id in active_ids if sl_id else False

            if tp_active and sl_active:
                continue

            if not tp_active and sl_active:
                try:
                    cancel_stop_order(account_id, sl_id)
                except Exception:
                    pass
                self.root.after(0, lambda t=trade: self.log(f"OCO: TP по {t.get('ticker')} исчез/сработал, SL снят."))
            elif not sl_active and tp_active:
                try:
                    cancel_stop_order(account_id, tp_id)
                except Exception:
                    pass
                self.root.after(0, lambda t=trade: self.log(f"OCO: SL по {t.get('ticker')} исчез/сработал, TP снят."))
            else:
                self.root.after(0, lambda t=trade: self.log(f"OCO: обе защитные заявки по {t.get('ticker')} отсутствуют."))

            try:
                after_portfolio = get_total_portfolio_value(account_id)
                pnl = None
                if trade.get("entry_portfolio"):
                    pnl = after_portfolio - Decimal(str(trade["entry_portfolio"]))
                append_journal_row(
                    datetime.now().strftime("%d.%m.%Y"),
                    after_portfolio,
                    pnl,
                    trade.get("ticker", ""),
                    f"OCO exit detected; account={self.account_short_label(account_id)}",
                )
            except Exception as exc:
                self.root.after(0, lambda e=exc: self.log(f"OCO: не смог записать журнал: {e}"))

            trades.remove(trade)
            changed = True

        if changed:
            state["active_trades"] = trades
            save_state(state)
            self.root.after(0, self.refresh_all_async)


def main():
    set_windows_app_user_model_id()
    root = tk.Tk()
    apply_window_icon(root)
    app = JTradeDarkApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
